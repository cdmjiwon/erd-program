from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.ws = None
    
    def generate(self, tables_info, output_path='table_definition.xlsx'):
        self.wb.remove(self.wb.active)
        
        summary_sheet = self.wb.create_sheet("테이블 목록")
        self._create_summary_sheet(summary_sheet, tables_info)
        
        for table_name, table_info in tables_info.items():
            sheet = self.wb.create_sheet(table_name[:31])
            self._create_table_sheet(sheet, table_name, table_info)
        
        self.wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, sheet, tables_info):
        headers = ["테이블명", "컬럼 수", "PK 컬럼", "FK 개수"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        for table_name, table_info in tables_info.items():
            sheet.cell(row=row, column=1, value=table_name)
            sheet.cell(row=row, column=2, value=len(table_info['columns']))
            sheet.cell(row=row, column=3, value=", ".join(table_info['primary_keys']))
            sheet.cell(row=row, column=4, value=len(table_info['foreign_keys']))
            row += 1
        
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20
    
    def _create_table_sheet(self, sheet, table_name, table_info):
        sheet.merge_cells('A1:F1')
        title_cell = sheet.cell(row=1, column=1, value=f"테이블명: {table_name}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        headers = ["순번", "컬럼명", "데이터 타입", "NULL 허용", "기본값", "설명"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        row = 3
        for idx, col in enumerate(table_info['columns'], 1):
            col_name = col['name']
            col_type = str(col['type'])
            nullable = "Y" if col.get('nullable', True) else "N"
            default = str(col.get('default', '')) if col.get('default') is not None else ""
            
            pk_marker = " [PK]" if col_name in table_info['primary_keys'] else ""
            
            sheet.cell(row=row, column=1, value=idx).border = border
            sheet.cell(row=row, column=2, value=col_name + pk_marker).border = border
            sheet.cell(row=row, column=3, value=col_type).border = border
            sheet.cell(row=row, column=4, value=nullable).border = border
            sheet.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            sheet.cell(row=row, column=5, value=default).border = border
            sheet.cell(row=row, column=6, value="").border = border
            
            row += 1
        
        if table_info['foreign_keys']:
            row += 1
            sheet.merge_cells(f'A{row}:F{row}')
            fk_title = sheet.cell(row=row, column=1, value="외래키 정보")
            fk_title.font = Font(bold=True, size=12)
            fk_title.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            fk_title.alignment = Alignment(horizontal="center")
            
            row += 1
            fk_headers = ["외래키명", "컬럼", "참조 테이블", "참조 컬럼"]
            for col_idx, header in enumerate(fk_headers, 1):
                cell = sheet.cell(row=row, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="C5C5C5", end_color="C5C5C5", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            row += 1
            for fk in table_info['foreign_keys']:
                fk_name = fk.get('name', 'N/A')
                from_cols = ", ".join(fk['constrained_columns'])
                ref_table = fk['referred_table']
                to_cols = ", ".join(fk['referred_columns'])
                
                sheet.cell(row=row, column=1, value=fk_name).border = border
                sheet.cell(row=row, column=2, value=from_cols).border = border
                sheet.cell(row=row, column=3, value=ref_table).border = border
                sheet.cell(row=row, column=4, value=to_cols).border = border
                row += 1
        
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30

